#!/usr/bin/env python3
"""
Downloads Watcher for Spectre
Monitors ~/Downloads for new files and:
1. Detects Spectre inventory files by Excel content -> uploads to Spectre, deletes original
2. Sorts other work files (invoices, receipts, PDFs) into organized folders
3. Leaves personal/unrelated files alone
"""
import os
import sys
import time
import logging
import requests
import re
from pathlib import Path
from datetime import datetime
from typing import Optional, Tuple, Dict, Any

# Add backend to path for imports
SCRIPT_DIR = Path(__file__).parent
BACKEND_DIR = SCRIPT_DIR.parent / "backend"
sys.path.insert(0, str(BACKEND_DIR))

from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not available, Excel content detection disabled")

# Import backend's extraction function for consistent detection
try:
    from core.parsers import extract_header_metadata
    BACKEND_AVAILABLE = True
except ImportError:
    BACKEND_AVAILABLE = False
    print("Warning: Backend parsers not available, using fallback detection")

# Configuration
WATCH_DIR = Path(os.environ.get("SPECTRE_WATCH_DIR", os.path.expanduser("~/Downloads")))
API_URL = os.environ.get("SPECTRE_API_URL", "http://localhost:8000")
# Sort destinations
SORT_BASE = Path(os.environ.get("SPECTRE_SORT_BASE", os.path.expanduser("~/Documents/Sorted")))
SORT_FOLDERS = {
    "invoices": SORT_BASE / "Invoices",
    "receipts": SORT_BASE / "Receipts",
    "training": SORT_BASE / "Training",
    "reports": SORT_BASE / "Reports",
    "contracts": SORT_BASE / "Contracts",
    "other_work": SORT_BASE / "Other",
}

# File patterns to watch
EXCEL_EXTENSIONS = {'.xlsx', '.xls'}
PDF_EXTENSION = '.pdf'
CSV_EXTENSION = '.csv'
WORK_EXTENSIONS = EXCEL_EXTENSIONS | {PDF_EXTENSION, CSV_EXTENSION}

# Logging
LOG_DIR = Path(os.path.expanduser("~/.local/share/spectre"))
LOG_DIR.mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(LOG_DIR / "downloads_watcher.log")
    ]
)
logger = logging.getLogger(__name__)


def ensure_dirs():
    """Create necessary directories."""
    for folder in SORT_FOLDERS.values():
        folder.mkdir(parents=True, exist_ok=True)


def is_spectre_inventory(filepath: Path) -> Tuple[bool, Optional[Dict[str, Any]]]:
    """
    Check if an Excel file is a Spectre inventory file by examining content.
    Uses backend's extract_header_metadata for consistent detection.

    A file is considered a Spectre inventory if it has BOTH:
    - Date in row 1 (pattern: MM/DD/YYYY)
    - Site name in row 2 (text before first parenthesis)

    Returns (is_inventory, metadata_dict)
    """
    if filepath.suffix.lower() not in EXCEL_EXTENSIONS:
        return False, None

    # Use backend's extraction function if available (preferred)
    if BACKEND_AVAILABLE:
        try:
            metadata = extract_header_metadata(str(filepath))
            if metadata.get("extracted"):
                return True, metadata
            return False, None
        except Exception as e:
            logger.debug(f"Backend extraction failed, using fallback: {e}")

    # Fallback: manual extraction if backend not available
    if not OPENPYXL_AVAILABLE:
        return False, None

    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)

        # Find data sheet - prioritize "data" in name
        data_sheet = None
        for name in wb.sheetnames:
            if 'data' in name.lower():
                data_sheet = wb[name]
                break

        # Second: look for "inventory" but not "summary"
        if not data_sheet:
            for name in wb.sheetnames:
                lower = name.lower()
                if 'inventory' in lower and 'summary' not in lower:
                    data_sheet = wb[name]
                    break

        # Fallback to second sheet (often data) or first
        if not data_sheet:
            if len(wb.sheetnames) > 1:
                data_sheet = wb[wb.sheetnames[1]]
            else:
                data_sheet = wb.active

        if not data_sheet:
            wb.close()
            return False, None

        # Check row 1 for date (MUST have date)
        row1_val = data_sheet.cell(row=1, column=1).value
        date_match = None
        if row1_val and isinstance(row1_val, str):
            date_match = re.search(r'(\d{1,2})/(\d{1,2})/(\d{4})$', row1_val.strip())

        # Check row 2 for site name (MUST have site)
        row2_val = data_sheet.cell(row=2, column=1).value
        site_name = None
        if row2_val and isinstance(row2_val, str):
            site_match = re.match(r'^([^(]+)', row2_val.strip())
            if site_match:
                site_name = site_match.group(1).strip()

        wb.close()

        # STRICT: Must have BOTH date AND site name
        if not date_match or not site_name:
            return False, None

        # Build metadata
        month, day, year = date_match.groups()
        metadata = {
            "extracted": True,
            "inventory_date": f"{year}-{month.zfill(2)}-{day.zfill(2)}",
            "site_name": site_name,
            "site_id": re.sub(r'[^a-z0-9]+', '_', site_name.lower()).strip('_'),
        }

        return True, metadata

    except Exception as e:
        logger.debug(f"Error checking Excel content: {e}")
        return False, None


def categorize_file(filepath: Path) -> Optional[str]:
    """
    Categorize a file based on its name using pattern matching.
    Returns category: invoices, receipts, training, reports, contracts, other_work
    Returns None for files that don't match any work pattern.
    """
    filename = filepath.name.lower()

    if any(kw in filename for kw in ['invoice', 'inv_', 'inv-', 'billing', 'charge']):
        return "invoices"
    if any(kw in filename for kw in ['receipt', 'rcpt', 'payment', 'confirmation']):
        return "receipts"
    if any(kw in filename for kw in ['training', 'manual', 'guide', 'tutorial', 'instruction', 'sop', 'procedure']):
        return "training"
    if any(kw in filename for kw in ['report', 'summary', 'analysis', 'audit', 'review']):
        return "reports"
    if any(kw in filename for kw in ['contract', 'agreement', 'terms', 'nda']):
        return "contracts"
    if any(kw in filename for kw in ['order', 'purchase', 'vendor', 'inventory', 'menu', 'schedule', 'roster']):
        return "other_work"

    return "other_work"


def upload_to_spectre(filepath: Path, site_id: Optional[str] = None) -> Tuple[bool, bool]:
    """
    Upload a file to the Spectre API.
    Returns (success, is_duplicate) - if duplicate, we can still delete the file.
    """
    logger.info(f"Uploading to Spectre: {filepath.name}")

    try:
        with open(filepath, 'rb') as f:
            files = {'file': (filepath.name, f)}
            data = {}
            if site_id:
                data['site_id'] = site_id

            response = requests.post(
                f"{API_URL}/api/files/upload",
                files=files,
                data=data,  # Form data, not query params
                timeout=60
            )

        if response.status_code == 200:
            result = response.json()
            logger.info(f"Uploaded successfully: {filepath.name} -> {result.get('id', 'unknown')}")
            return True, False
        elif response.status_code == 400:
            # Check if it's a duplicate error
            try:
                error_detail = response.json().get("detail", "")
                if "duplicate" in error_detail.lower():
                    logger.info(f"File already in Spectre (duplicate): {filepath.name}")
                    return True, True  # Success = True because file is already there
            except:
                pass
            logger.error(f"Upload failed ({response.status_code}): {response.text}")
            return False, False
        else:
            logger.error(f"Upload failed ({response.status_code}): {response.text}")
            return False, False
    except requests.exceptions.ConnectionError:
        logger.error(f"Connection error - is the backend running at {API_URL}?")
        return False, False
    except Exception as e:
        logger.error(f"Upload error: {e}")
        return False, False


def move_file(filepath: Path, dest_dir: Path) -> bool:
    """Move file to destination directory."""
    try:
        dest_dir.mkdir(parents=True, exist_ok=True)
        dest = dest_dir / filepath.name

        # Handle duplicate names
        if dest.exists():
            base = filepath.stem
            ext = filepath.suffix
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            dest = dest_dir / f"{base}_{timestamp}{ext}"

        filepath.rename(dest)
        logger.info(f"Moved to: {dest}")
        return True
    except Exception as e:
        logger.error(f"Move failed: {e}")
        return False


def delete_file(filepath: Path) -> bool:
    """Delete a file."""
    try:
        filepath.unlink()
        logger.info(f"Deleted: {filepath.name}")
        return True
    except Exception as e:
        logger.error(f"Delete failed: {e}")
        return False


def process_file(filepath: Path):
    """Process a single file: detect type and route appropriately."""
    if not filepath.exists():
        return

    ext = filepath.suffix.lower()

    # Skip non-work files
    if ext not in WORK_EXTENSIONS:
        logger.debug(f"Ignoring non-work file: {filepath.name}")
        return

    # Wait for file to be fully written
    time.sleep(2)

    if not filepath.exists():
        return

    logger.info(f"Processing: {filepath.name}")

    # Check if it's a Spectre inventory file (Excel only)
    if ext in EXCEL_EXTENSIONS:
        is_inventory, metadata = is_spectre_inventory(filepath)

        if is_inventory:
            logger.info(f"Detected Spectre inventory file: {filepath.name}")
            if metadata:
                logger.info(f"  Site: {metadata.get('site_name', 'unknown')}, Date: {metadata.get('inventory_date', 'unknown')}")

            # Upload to Spectre
            site_id = metadata.get('site_id') if metadata else None
            success, is_duplicate = upload_to_spectre(filepath, site_id)
            if success:
                # Delete from Downloads on success (including duplicates)
                if is_duplicate:
                    logger.info(f"Removing duplicate from Downloads: {filepath.name}")
                delete_file(filepath)
            return

    # Not a Spectre file - categorize and sort
    category = categorize_file(filepath)

    if category is None:
        logger.info(f"Skipping personal file: {filepath.name}")
        return

    if category in SORT_FOLDERS:
        dest_dir = SORT_FOLDERS[category]
        logger.info(f"Categorized as '{category}': {filepath.name}")
        move_file(filepath, dest_dir)
    else:
        logger.info(f"Unknown category '{category}', leaving in place: {filepath.name}")


class DownloadsHandler(FileSystemEventHandler):
    """Handle file system events in Downloads folder."""

    def __init__(self):
        self.recently_processed = {}  # Track to avoid duplicates

    def _should_process(self, path: str) -> bool:
        """Check if we should process this file."""
        # Avoid processing the same file multiple times in quick succession
        now = time.time()
        if path in self.recently_processed:
            if now - self.recently_processed[path] < 5:
                return False
        self.recently_processed[path] = now

        # Clean old entries
        self.recently_processed = {
            k: v for k, v in self.recently_processed.items()
            if now - v < 60
        }
        return True

    def on_created(self, event):
        if event.is_directory:
            return
        if not self._should_process(event.src_path):
            return
        process_file(Path(event.src_path))

    def on_moved(self, event):
        if event.is_directory:
            return
        if not self._should_process(event.dest_path):
            return
        process_file(Path(event.dest_path))


def process_existing():
    """Process any matching files already in Downloads."""
    logger.info(f"Checking for existing files in {WATCH_DIR}")
    for filepath in WATCH_DIR.iterdir():
        if filepath.is_file() and filepath.suffix.lower() in WORK_EXTENSIONS:
            logger.info(f"Found existing file: {filepath.name}")
            process_file(filepath)


def main():
    ensure_dirs()

    print(f"""
╔══════════════════════════════════════════════════════════════╗
║           Spectre Downloads Watcher                         ║
╠══════════════════════════════════════════════════════════════╣
║  Watching: {str(WATCH_DIR):<48} ║
║  API URL:  {API_URL:<48} ║
╠══════════════════════════════════════════════════════════════╣
║  Spectre inventory files -> Upload & delete                 ║
║  Work files (invoices, receipts, etc.) -> Sort to folders   ║
║  Personal files -> Leave alone                              ║
╚══════════════════════════════════════════════════════════════╝
    """)

    logger.info("Sort destinations:")
    for cat, path in SORT_FOLDERS.items():
        logger.info(f"  {cat}: {path}")

    # Process existing files first
    process_existing()

    # Start watching
    event_handler = DownloadsHandler()
    observer = Observer()
    observer.schedule(event_handler, str(WATCH_DIR), recursive=False)
    observer.start()

    logger.info("Watching for new files... (Ctrl+C to stop)")

    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        logger.info("Stopping watcher...")
        observer.stop()

    observer.join()
    logger.info("Watcher stopped.")


if __name__ == "__main__":
    main()

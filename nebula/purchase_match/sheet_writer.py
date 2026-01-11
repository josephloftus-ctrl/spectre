"""
Clean XLSX Writer - Modify Excel files while preserving formatting.

Handles sorting, filtering, and updating count sheets without
disrupting template styles, formulas, or structure.
"""

import logging
from pathlib import Path
from typing import Optional, List, Dict, Any, Literal
from copy import copy
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell

logger = logging.getLogger(__name__)

SortOrder = Literal["description", "sku", "category", "vendor", "price"]


def load_template(template_path: Path) -> Optional[Any]:
    """
    Load an Excel template while preserving all formatting.

    Returns the workbook object or None if failed.
    """
    try:
        # data_only=False preserves formulas
        # keep_vba=False since we don't have macros
        wb = load_workbook(template_path, data_only=False)
        return wb
    except Exception as e:
        logger.error(f"Failed to load template {template_path}: {e}")
        return None


def find_header_row(ws: Worksheet, max_rows: int = 10) -> Optional[int]:
    """
    Find the header row by looking for common column names.

    Returns 1-indexed row number or None.
    """
    header_keywords = {
        "description", "item", "sku", "quantity", "qty",
        "price", "total", "vendor", "category", "uom"
    }

    for row_num in range(1, max_rows + 1):
        row_values = []
        for cell in ws[row_num]:
            if cell.value:
                row_values.append(str(cell.value).lower().strip())

        # Check if this row has header-like values
        matches = sum(1 for v in row_values if any(kw in v for kw in header_keywords))
        if matches >= 2:  # At least 2 header-like columns
            return row_num

    return None


def find_column_index(ws: Worksheet, header_row: int, column_name: str) -> Optional[int]:
    """
    Find column index (1-indexed) by header name.

    Searches for partial matches (e.g., "desc" matches "Item Description").
    """
    search_term = column_name.lower()

    for col_idx, cell in enumerate(ws[header_row], start=1):
        if cell.value:
            header = str(cell.value).lower()
            if search_term in header:
                return col_idx

    return None


def copy_cell_style(source: Cell, target: Cell):
    """Copy all styling from source cell to target cell."""
    if source.has_style:
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)


def sort_worksheet(
    ws: Worksheet,
    sort_by: SortOrder = "description",
    header_row: Optional[int] = None
) -> bool:
    """
    Sort worksheet data rows while preserving formatting.

    Args:
        ws: Worksheet to sort
        sort_by: Column to sort by
        header_row: Header row number (auto-detected if None)

    Returns:
        True if successful
    """
    # Find header row
    if header_row is None:
        header_row = find_header_row(ws)
        if header_row is None:
            logger.warning("Could not find header row")
            return False

    # Map sort_by to column search terms
    column_map = {
        "description": "description",
        "sku": "sku",
        "category": "category",
        "vendor": "vendor",
        "price": "price"
    }

    search_term = column_map.get(sort_by, "description")
    sort_col = find_column_index(ws, header_row, search_term)

    if sort_col is None:
        # Fallback: try item for description
        if sort_by == "description":
            sort_col = find_column_index(ws, header_row, "item")

    if sort_col is None:
        logger.warning(f"Could not find column for sorting by '{sort_by}'")
        return False

    # Collect data rows (after header)
    data_start = header_row + 1
    data_rows = []

    for row_num in range(data_start, ws.max_row + 1):
        row_data = []
        row_styles = []
        has_data = False

        for col_num in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_num, column=col_num)
            row_data.append(cell.value)

            # Store style info
            style_info = {
                "font": copy(cell.font) if cell.has_style else None,
                "fill": copy(cell.fill) if cell.has_style else None,
                "border": copy(cell.border) if cell.has_style else None,
                "alignment": copy(cell.alignment) if cell.has_style else None,
                "number_format": cell.number_format,
            }
            row_styles.append(style_info)

            if cell.value is not None and str(cell.value).strip():
                has_data = True

        if has_data:
            sort_value = row_data[sort_col - 1] if sort_col <= len(row_data) else ""
            data_rows.append({
                "sort_key": str(sort_value or "").lower(),
                "data": row_data,
                "styles": row_styles
            })

    if not data_rows:
        logger.info("No data rows to sort")
        return True

    # Sort rows
    data_rows.sort(key=lambda r: r["sort_key"])

    # Write sorted rows back
    for i, row_info in enumerate(data_rows):
        row_num = data_start + i

        for col_num, (value, style) in enumerate(zip(row_info["data"], row_info["styles"]), start=1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value

            # Restore style
            if style["font"]:
                cell.font = style["font"]
            if style["fill"]:
                cell.fill = style["fill"]
            if style["border"]:
                cell.border = style["border"]
            if style["alignment"]:
                cell.alignment = style["alignment"]
            if style["number_format"]:
                cell.number_format = style["number_format"]

    logger.info(f"Sorted {len(data_rows)} rows by '{sort_by}'")
    return True


def generate_sorted_template(
    template_path: Path,
    sort_by: SortOrder = "description"
) -> Optional[bytes]:
    """
    Generate a sorted copy of a template.

    Args:
        template_path: Path to the template file
        sort_by: Column to sort by

    Returns:
        Bytes of the sorted Excel file, or None if failed
    """
    wb = load_template(template_path)
    if not wb:
        return None

    # Sort the active sheet (or first sheet)
    ws = wb.active
    if not sort_worksheet(ws, sort_by):
        logger.warning("Sorting failed, returning unsorted template")

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output.read()


def update_template_items(
    template_path: Path,
    items: List[Dict[str, Any]],
    sort_by: Optional[SortOrder] = None
) -> Optional[bytes]:
    """
    Update template with new item data while preserving formatting.

    Args:
        template_path: Path to the template
        items: List of items with sku, description, quantity, price, etc.
        sort_by: Optional column to sort by after update

    Returns:
        Bytes of the updated Excel file
    """
    wb = load_template(template_path)
    if not wb:
        return None

    ws = wb.active
    header_row = find_header_row(ws)

    if header_row is None:
        logger.error("Could not find header row in template")
        return None

    # Find relevant columns
    col_map = {}
    for col_name in ["sku", "description", "item", "quantity", "qty", "price", "vendor"]:
        idx = find_column_index(ws, header_row, col_name)
        if idx:
            col_map[col_name] = idx

    # Get a sample row for formatting reference
    sample_row = header_row + 1
    sample_styles = {}
    for col_num in range(1, ws.max_column + 1):
        cell = ws.cell(row=sample_row, column=col_num)
        if cell.has_style:
            sample_styles[col_num] = {
                "font": copy(cell.font),
                "fill": copy(cell.fill),
                "border": copy(cell.border),
                "alignment": copy(cell.alignment),
                "number_format": cell.number_format,
            }

    # Clear existing data rows
    for row_num in range(header_row + 1, ws.max_row + 1):
        for col_num in range(1, ws.max_column + 1):
            ws.cell(row=row_num, column=col_num).value = None

    # Write new items
    for i, item in enumerate(items):
        row_num = header_row + 1 + i

        # Write values based on column mapping
        if "sku" in col_map and "sku" in item:
            ws.cell(row=row_num, column=col_map["sku"]).value = item["sku"]

        desc_col = col_map.get("description") or col_map.get("item")
        if desc_col and "description" in item:
            ws.cell(row=row_num, column=desc_col).value = item["description"]

        qty_col = col_map.get("quantity") or col_map.get("qty")
        if qty_col and "quantity" in item:
            ws.cell(row=row_num, column=qty_col).value = item["quantity"]

        if "price" in col_map and "price" in item:
            ws.cell(row=row_num, column=col_map["price"]).value = item["price"]

        if "vendor" in col_map and "vendor" in item:
            ws.cell(row=row_num, column=col_map["vendor"]).value = item["vendor"]

        # Apply sample formatting
        for col_num, style in sample_styles.items():
            cell = ws.cell(row=row_num, column=col_num)
            if style.get("font"):
                cell.font = style["font"]
            if style.get("fill"):
                cell.fill = style["fill"]
            if style.get("border"):
                cell.border = style["border"]
            if style.get("alignment"):
                cell.alignment = style["alignment"]
            if style.get("number_format"):
                cell.number_format = style["number_format"]

    # Optionally sort
    if sort_by:
        sort_worksheet(ws, sort_by, header_row)

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output.read()

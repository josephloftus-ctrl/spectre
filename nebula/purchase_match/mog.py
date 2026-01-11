"""
Master Order Guide (MOG) Parser and Index.

Loads vendor order guides to:
1. Validate if a SKU exists in the catalog
2. Find similar items by description for SKU correction suggestions
"""

import openpyxl
from pathlib import Path
from dataclasses import dataclass
from typing import Optional, Dict, List, Set
from decimal import Decimal
import re


@dataclass
class MOGItem:
    """A single item from a Master Order Guide."""
    sku: str
    description: str
    vendor: str  # sysco, gfs, vistar_east, vistar_midatlantic
    uom: str
    price: Optional[Decimal]
    brand: Optional[str]
    category: Optional[str]


class MOGIndex:
    """
    Index of all items from Master Order Guides.

    Provides fast lookup by SKU and description similarity search.
    """

    def __init__(self):
        self._items: Dict[str, MOGItem] = {}  # SKU -> item
        self._by_vendor: Dict[str, Set[str]] = {}  # vendor -> set of SKUs
        self._descriptions: Dict[str, str] = {}  # normalized description -> SKU

    @property
    def total_items(self) -> int:
        return len(self._items)

    @property
    def vendors(self) -> List[str]:
        return list(self._by_vendor.keys())

    def add_item(self, item: MOGItem):
        """Add an item to the index."""
        self._items[item.sku] = item

        if item.vendor not in self._by_vendor:
            self._by_vendor[item.vendor] = set()
        self._by_vendor[item.vendor].add(item.sku)

        # Index by normalized description for similarity search
        norm_desc = self._normalize_description(item.description)
        if norm_desc:
            self._descriptions[norm_desc] = item.sku

    def lookup_sku(self, sku: str) -> Optional[MOGItem]:
        """Check if a SKU exists in any order guide."""
        return self._items.get(sku)

    def sku_exists(self, sku: str) -> bool:
        """Quick check if SKU exists."""
        return sku in self._items

    def get_vendor_items(self, vendor: str) -> Set[str]:
        """Get all SKUs for a vendor."""
        return self._by_vendor.get(vendor, set())

    def all_items(self) -> List[MOGItem]:
        """Get all items in the index."""
        return list(self._items.values())

    def find_by_description(self, description: str, limit: int = 5) -> List[MOGItem]:
        """
        Find items with similar descriptions.

        Uses simple word overlap for now - can be enhanced with fuzzy matching.
        """
        if not description:
            return []

        query_words = set(self._normalize_description(description).split())
        if not query_words:
            return []

        # Score all items by word overlap
        scores = []
        for norm_desc, sku in self._descriptions.items():
            item_words = set(norm_desc.split())
            overlap = len(query_words & item_words)
            if overlap > 0:
                # Jaccard similarity
                similarity = overlap / len(query_words | item_words)
                scores.append((similarity, sku))

        # Sort by score descending
        scores.sort(reverse=True, key=lambda x: x[0])

        return [self._items[sku] for _, sku in scores[:limit]]

    def _normalize_description(self, desc: str) -> str:
        """Normalize description for comparison."""
        if not desc:
            return ""
        # Lowercase, remove punctuation, collapse whitespace
        desc = desc.lower()
        desc = re.sub(r'[^\w\s]', ' ', desc)
        desc = re.sub(r'\s+', ' ', desc).strip()
        return desc


def detect_vendor(filename: str) -> str:
    """Detect vendor from filename."""
    name = filename.lower()
    if 'sysco' in name:
        return 'sysco'
    elif 'gfs' in name:
        return 'gfs'
    elif 'vistar' in name:
        if 'east' in name:
            return 'vistar_east'
        elif 'mid' in name or 'atlantic' in name:
            return 'vistar_midatlantic'
        return 'vistar'
    return 'unknown'


def parse_mog_file(filepath: Path) -> List[MOGItem]:
    """
    Parse a single MOG Excel file.

    Expected format:
    - Row 6: Headers
    - Row 7+: Data
    - Column B: Dist # (SKU)
    - Column C: Item Description
    - Column G: UOM
    - Column H: Price
    - Column K: Category
    - Column L: Brand
    """
    items = []
    vendor = detect_vendor(filepath.name)

    try:
        # Don't use read_only=True - these files have unusual structure
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active

        # Find header row (look for "Dist #")
        header_row = None
        for row_num in range(1, 15):
            cell_b = ws.cell(row=row_num, column=2).value
            if cell_b and 'dist' in str(cell_b).lower():
                header_row = row_num
                break

        if header_row is None:
            # Default to row 6
            header_row = 6

        # Parse data rows - read actual max_row
        max_row = ws.max_row
        for row_num in range(header_row + 1, max_row + 1):
            row = [ws.cell(row=row_num, column=c).value for c in range(1, 22)]
            if len(row) < 12:
                continue

            sku = row[1]  # Column B
            description = row[2]  # Column C

            if not sku or not description:
                continue

            # Clean SKU
            sku = str(sku).strip()
            if not sku:
                continue

            # Parse price
            price = None
            if row[7]:  # Column H
                price_str = str(row[7]).replace('$', '').replace(',', '').strip()
                try:
                    price = Decimal(price_str)
                except:
                    pass

            items.append(MOGItem(
                sku=sku,
                description=str(description).strip(),
                vendor=vendor,
                uom=str(row[6]).strip() if row[6] else '',
                price=price,
                brand=str(row[11]).strip() if row[11] else None,
                category=str(row[10]).strip() if row[10] else None,
            ))

        wb.close()

    except Exception as e:
        print(f"Error parsing {filepath}: {e}")

    return items


def load_mog_directory(mog_dir: Path) -> MOGIndex:
    """
    Load all MOG files from a directory into an index.

    Args:
        mog_dir: Path to directory containing MOG Excel files

    Returns:
        MOGIndex with all items loaded
    """
    index = MOGIndex()

    if not mog_dir.exists():
        return index

    for filepath in mog_dir.glob("*.xlsx"):
        items = parse_mog_file(filepath)
        for item in items:
            index.add_item(item)
        print(f"Loaded {len(items)} items from {filepath.name}")

    print(f"Total MOG items: {index.total_items}")
    return index

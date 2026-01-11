"""
Canon Loader - Parse OrderMaestro IPS exports into PurchaseRecords.

The IPS export is the source of truth for "what was actually purchased."
Multi-month files get merged into a single canon.
"""

from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Optional
from datetime import date
import re

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .models import PurchaseRecord
from .config import Config, normalize_vendor


def _decode_xml_value(value: str) -> str:
    """
    Decode XML-encoded values from Excel.

    Examples:
        _x0020_ -> space
        _x0031_ -> "1" (ASCII 49)
        Farmer_x0020_Brothers_x0020_Co -> Farmer Brothers Co
        _x0031_17379 -> 117379
    """
    if not value or "_x" not in value:
        return value

    def replace_code(match):
        hex_code = match.group(1)
        try:
            return chr(int(hex_code, 16))
        except ValueError:
            return match.group(0)

    return re.sub(r'_x([0-9A-Fa-f]{4})_', replace_code, value)


# Column mappings (0-indexed)
# IPS Format: Headers on row 10, data starts row 11
HEADER_ROW = 10
DATA_START_ROW = 11

# Expected columns (may have encoded names like Item_x0020_Number)
COLUMN_PATTERNS = {
    "distributor": ["Distributor", "distributor"],
    "item_number": ["Item Number", "Item_x0020_Number", "item_number", "ItemNumber"],
    "description": ["Description", "description", "Item Description", "Item_x0020_Description"],
    "brand": ["Brand", "brand"],
    "uom": ["Unit of Measure", "Unit_x0020_of_x0020_Measure", "UOM", "uom"],
    "pack": ["Pack", "pack"],
    "price": ["Invoiced Item Price", "Invoiced_x0020_Item_x0020_Price", "Price", "price"],
}


def _find_column_index(headers: list[str], patterns: list[str]) -> Optional[int]:
    """Find column index matching any of the patterns (case-insensitive)."""
    for i, header in enumerate(headers):
        if header is None:
            continue
        header_lower = str(header).lower().strip()
        for pattern in patterns:
            if pattern.lower() == header_lower:
                return i
    return None


def _parse_price(value) -> Optional[Decimal]:
    """Parse price value to Decimal, handling various formats."""
    if value is None:
        return None

    # Handle numeric values directly
    if isinstance(value, (int, float)):
        return Decimal(str(round(value, 2)))

    # Handle string values
    str_value = str(value).strip()
    if not str_value:
        return None

    # Remove currency symbols and commas
    str_value = re.sub(r'[$,]', '', str_value)

    try:
        return Decimal(str_value).quantize(Decimal("0.01"))
    except InvalidOperation:
        return None


def _find_data_sheet(workbook) -> Optional[Worksheet]:
    """Find the sheet containing purchase data."""
    # Look for sheet named "Purchasing Details Raw Data"
    target_names = [
        "Purchasing Details Raw Data",
        "Purchasing_Details_Raw_Data",
        "Raw Data",
    ]

    for name in target_names:
        if name in workbook.sheetnames:
            return workbook[name]

    # Fallback: use sheet with most rows
    max_rows = 0
    best_sheet = None
    for sheet in workbook.worksheets:
        if sheet.max_row > max_rows:
            max_rows = sheet.max_row
            best_sheet = sheet

    return best_sheet


def load_canon(
    file_paths: list[str | Path],
    config: Config,
    header_row: int = HEADER_ROW,
    data_start_row: int = DATA_START_ROW,
) -> list[PurchaseRecord]:
    """
    Load purchase records from one or more IPS export files.

    Args:
        file_paths: Paths to XLSX files
        config: Config with vendor aliases for normalization
        header_row: Row number containing headers (1-indexed, default 10)
        data_start_row: Row number where data starts (1-indexed, default 11)

    Returns:
        List of PurchaseRecord objects from all files
    """
    all_records = []

    for file_path in file_paths:
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"IPS file not found: {path}")

        records = _load_single_file(path, config, header_row, data_start_row)
        all_records.extend(records)

    return all_records


def _load_single_file(
    path: Path,
    config: Config,
    header_row: int,
    data_start_row: int,
) -> list[PurchaseRecord]:
    """Load records from a single IPS file."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = _find_data_sheet(workbook)

    if sheet is None:
        raise ValueError(f"No data sheet found in {path}")

    # Get headers from specified row
    headers = []
    for cell in sheet[header_row]:
        headers.append(cell.value)

    # Find column indices
    col_idx = {}
    for field, patterns in COLUMN_PATTERNS.items():
        idx = _find_column_index(headers, patterns)
        col_idx[field] = idx

    # Verify required columns exist
    required = ["item_number", "distributor", "price"]
    missing = [f for f in required if col_idx.get(f) is None]
    if missing:
        raise ValueError(f"Missing required columns in {path}: {missing}")

    records = []
    for row_num, row in enumerate(sheet.iter_rows(min_row=data_start_row), start=data_start_row):
        # Get cell values by index, decoding XML-encoded values
        def get_val(field):
            idx = col_idx.get(field)
            if idx is None or idx >= len(row):
                return None
            val = row[idx].value
            if isinstance(val, str):
                return _decode_xml_value(val)
            return val

        sku = get_val("item_number")
        if not sku:
            continue  # Skip rows without item number

        # Decode and normalize vendor
        vendor_raw = get_val("distributor") or ""
        if isinstance(vendor_raw, str):
            vendor_raw = _decode_xml_value(vendor_raw)
        vendor = normalize_vendor(vendor_raw, config)

        price = _parse_price(get_val("price"))
        if price is None:
            continue  # Skip rows without valid price

        # Decode SKU in case it has encoded leading digits
        sku_str = str(sku).strip()
        if isinstance(sku, str):
            sku_str = _decode_xml_value(sku_str)

        record = PurchaseRecord(
            sku=sku_str,
            vendor=vendor or "unknown",
            vendor_raw=str(vendor_raw).strip(),
            price=price,
            description=str(get_val("description") or "").strip(),
            brand=str(get_val("brand") or "").strip() or None,
            uom=str(get_val("uom") or "").strip() or None,
            pack=str(get_val("pack") or "").strip() or None,
        )
        records.append(record)

    workbook.close()
    return records

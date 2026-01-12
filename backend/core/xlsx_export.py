"""
XLSX Export Module - OrderMaestro Compatible Exports

Generates Excel files matching OrderMaestro's exact template formats.

IMPORTANT: MyOrders/OrderMaestro uses DIFFERENT formats for downloads vs uploads:

DOWNLOAD formats (what we parse from OrderMaestro):
- Valuation Report: 3 sheets, 14 columns, header rows 1-8, data at row 10+

UPLOAD formats (what we generate FOR OrderMaestro):
- Inventory Upload Template: 1 sheet, 23 columns, headers row 1, data row 2+
- Shopping Cart Template: 1 sheet, 4 columns (Dist #, GTIN, Quantity, Break Quantity)
- Shopping List Template: 1 sheet, 3 columns (Dist #, GTIN, Cust #)

WARNING from training docs: "if you alter this template in any way, it will not upload"
"""

from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


# =============================================================================
# UPLOAD TEMPLATE COLUMNS (for uploading TO OrderMaestro)
# =============================================================================

# Inventory Upload Template - 23 columns (from training docs)
INVENTORY_UPLOAD_COLUMNS = [
    "Item Description",
    "Dist #",           # Required
    "Cust #",           # Required for off-catalog items
    "Quantity",
    "Break Quantity",
    "UOM",
    "Break Uom",
    "Location",         # Storage location name
    "Area",             # Sub-location
    "Place",            # Secondary sub-location
    "Distribution Center",
    "Brand",
    "Mfg",
    "Mfg #",
    "Pack",
    "GTIN",
    "Price",
    "Break Price",
    "Distributor",
    "Upc",
    "Catch Weight",
    "Average Weight",
    "Units Per Case",
]

# Shopping Cart Upload Template - 4 columns (from training docs)
CART_UPLOAD_COLUMNS = [
    "Dist #",           # Required - item number
    "GTIN",             # Optional
    "Quantity",         # Required - order quantity
    "Break Quantity",   # Optional - for split case orders
]

# Shopping List Upload Template - 3 columns (from training docs)
SHOPPING_LIST_UPLOAD_COLUMNS = [
    "Dist #",           # Required - item number
    "GTIN",             # Optional
    "Cust #",           # Optional
]


# =============================================================================
# VALUATION REPORT COLUMNS (for parsing downloads FROM OrderMaestro)
# =============================================================================

VALUATION_REPORT_COLUMNS = [
    "Compass Group USA->GL Codes",
    "Dist #",
    "Item Description",
    "Pack",
    "Quantity",
    "UOM",
    "Unit Price",
    "Price Code",
    "Total Price",
    "Distributor",
    "DC Name",
    "DC Category",
    "GTIN",
    "Brand",
]


# =============================================================================
# INVENTORY UPLOAD TEMPLATE (for uploading counts back to OrderMaestro)
# =============================================================================

def create_inventory_upload_workbook(
    items: List[Dict[str, Any]],
) -> BytesIO:
    """
    Create an Inventory Upload Template for OrderMaestro.

    This is the format required to upload inventory counts back to OrderMaestro.
    Single sheet, 23 columns, headers in row 1, data starting row 2.

    Args:
        items: List of inventory items with keys matching column names

    Returns:
        BytesIO buffer containing the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Upload"

    # Header row (row 1) - must be exact column names
    ws.append(INVENTORY_UPLOAD_COLUMNS)

    # Bold the header row
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Data rows starting at row 2
    for item in items:
        row_data = extract_inventory_upload_row(item)
        ws.append(row_data)

    # Set reasonable column widths
    column_widths = [35, 12, 12, 10, 12, 8, 10, 20, 15, 15, 20, 15, 15, 12, 12, 18, 10, 10, 20, 15, 12, 12, 12]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def extract_inventory_upload_row(item: Dict[str, Any]) -> List[Any]:
    """Extract a row for inventory upload template."""
    def get_val(keys: List[str], default: Any = "") -> Any:
        for key in keys:
            if key in item and item[key] is not None:
                return item[key]
            # Case-insensitive match
            for k, v in item.items():
                if k and k.lower() == key.lower() and v is not None:
                    return v
        return default

    return [
        get_val(["Item Description", "description", "item_name"]),
        get_val(["Dist #", "dist_num", "sku", "item_number"]),
        get_val(["Cust #", "cust_num", "customer_number"]),
        get_val(["Quantity", "quantity", "qty", "counted_qty"], 0),
        get_val(["Break Quantity", "break_quantity", "break_qty"]),
        get_val(["UOM", "uom", "unit_of_measure"]),
        get_val(["Break Uom", "break_uom"]),
        get_val(["Location", "location", "storage_location"]),
        get_val(["Area", "area", "sub_location"]),
        get_val(["Place", "place", "secondary_location"]),
        get_val(["Distribution Center", "DC Name", "dc_name"]),
        get_val(["Brand", "brand"]),
        get_val(["Mfg", "mfg", "manufacturer"]),
        get_val(["Mfg #", "mfg_num"]),
        get_val(["Pack", "pack", "pack_size"]),
        get_val(["GTIN", "gtin", "upc", "barcode"]),
        get_val(["Price", "Unit Price", "unit_price", "price"]),
        get_val(["Break Price", "break_price"]),
        get_val(["Distributor", "distributor", "vendor"]),
        get_val(["Upc", "upc"]),
        get_val(["Catch Weight", "catch_weight"]),
        get_val(["Average Weight", "average_weight"]),
        get_val(["Units Per Case", "units_per_case"]),
    ]


# =============================================================================
# SHOPPING CART UPLOAD TEMPLATE (for uploading orders to OrderMaestro)
# =============================================================================

def create_cart_upload_workbook(
    items: List[Dict[str, Any]],
) -> BytesIO:
    """
    Create a Shopping Cart Upload Template for OrderMaestro.

    This is the format required to upload shopping cart orders to OrderMaestro.
    Single sheet, 4 columns: Dist #, GTIN, Quantity, Break Quantity
    Only Dist # and Quantity are required.

    Args:
        items: List of cart items with sku and quantity

    Returns:
        BytesIO buffer containing the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Shopping Cart Upload"

    # Header row
    ws.append(CART_UPLOAD_COLUMNS)

    # Bold the header row
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Data rows
    for item in items:
        row_data = [
            item.get("Dist #") or item.get("sku") or item.get("dist_num") or "",
            item.get("GTIN") or item.get("gtin") or "",
            item.get("Quantity") or item.get("quantity") or 0,
            item.get("Break Quantity") or item.get("break_quantity") or "",
        ]
        ws.append(row_data)

    # Column widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 15

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# =============================================================================
# SHOPPING LIST UPLOAD TEMPLATE
# =============================================================================

def create_shopping_list_upload_workbook(
    items: List[Dict[str, Any]],
) -> BytesIO:
    """
    Create a Shopping List Upload Template for OrderMaestro.

    Single sheet, 3 columns: Dist #, GTIN, Cust #
    Only Dist # (item number) is required.

    Args:
        items: List of items with sku/dist_num

    Returns:
        BytesIO buffer containing the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Shopping List Upload"

    # Header row
    ws.append(SHOPPING_LIST_UPLOAD_COLUMNS)

    # Bold header
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Data rows - only need item numbers
    for item in items:
        row_data = [
            item.get("Dist #") or item.get("sku") or item.get("dist_num") or "",
            item.get("GTIN") or item.get("gtin") or "",
            item.get("Cust #") or item.get("cust_num") or "",
        ]
        ws.append(row_data)

    # Column widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 15

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# =============================================================================
# VALUATION REPORT FORMAT (for record-keeping/archiving, matches download format)
# =============================================================================

def create_valuation_report_workbook(
    site_name: str,
    items: List[Dict[str, Any]],
    printed_by: str = "Spectre",
) -> BytesIO:
    """
    Create a Valuation Report format workbook (matches OrderMaestro downloads).

    This format is for record-keeping and archiving. It matches the format
    that OrderMaestro produces when you download a valuation report.

    Structure:
    - Sheet 1: "Summary By Compass Group USA->G" (empty)
    - Sheet 2: "Multiple GL Codes" (headers only)
    - Sheet 3: "Data for Compass Group USA->GL " (main data)

    Args:
        site_name: Display name of the site
        items: List of inventory items
        printed_by: Name for "Printed By" field

    Returns:
        BytesIO buffer containing the Excel file
    """
    wb = Workbook()

    # Sheet 1: Summary (empty)
    ws_summary = wb.active
    ws_summary.title = "Summary By Compass Group USA->G"

    # Sheet 2: Multiple GL Codes (headers only)
    ws_gl = wb.create_sheet("Multiple GL Codes")
    ws_gl["A1"] = "Dist #"
    ws_gl["B1"] = "GL Codes"
    ws_gl["C1"] = "Item Description"

    # Sheet 3: Data (main content)
    ws_data = wb.create_sheet("Data for Compass Group USA->GL ")

    # Header metadata rows
    ws_data["A1"] = "Inventory Valuation Report from Inventory Management"
    ws_data["A2"] = f"{site_name} (COMPASS)"
    ws_data["A3"] = "Property of Compass Group USA Proprietary and Confidential"
    ws_data["A4"] = "Current Inventory"
    ws_data["A5"] = "Preferred Price Latest Invoice Price"
    ws_data["A6"] = f"Printed By: {printed_by}"
    # Rows 7-8 empty

    # Column headers at row 9
    for col_idx, header in enumerate(VALUATION_REPORT_COLUMNS, 1):
        ws_data.cell(row=9, column=col_idx, value=header)

    # Data rows starting at row 10
    for row_idx, item in enumerate(items, 10):
        row_data = extract_valuation_row(item)
        for col_idx, value in enumerate(row_data, 1):
            ws_data.cell(row=row_idx, column=col_idx, value=value)

    # Column widths
    widths = [30, 12, 40, 12, 10, 8, 10, 10, 12, 20, 25, 35, 18, 15]
    for col_idx, width in enumerate(widths, 1):
        ws_data.column_dimensions[get_column_letter(col_idx)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def extract_valuation_row(item: Dict[str, Any]) -> List[Any]:
    """Extract a row for valuation report format."""
    def get_val(keys: List[str], default: Any = "") -> Any:
        for key in keys:
            if key in item and item[key] is not None:
                return item[key]
            for k, v in item.items():
                if k and k.lower() == key.lower() and v is not None:
                    return v
        return default

    quantity = get_val(["Quantity", "quantity", "qty", "counted_qty"], 0)
    unit_price = get_val(["Unit Price", "unit_price", "price"])
    total_price = get_val(["Total Price", "total_price", "total"])

    # Calculate total if not provided
    if not total_price and quantity and unit_price:
        try:
            total_price = float(quantity) * float(unit_price)
        except (ValueError, TypeError):
            pass

    return [
        get_val(["Compass Group USA->GL Codes", "gl_codes", "GL Codes", "location"]),
        get_val(["Dist #", "dist_num", "sku", "item_number"]),
        get_val(["Item Description", "description", "item_name"]),
        get_val(["Pack", "pack", "pack_size"]),
        quantity,
        get_val(["UOM", "uom", "unit_of_measure"]),
        unit_price,
        get_val(["Price Code", "price_code"]),
        total_price,
        get_val(["Distributor", "distributor", "vendor"]),
        get_val(["DC Name", "dc_name"]),
        get_val(["DC Category", "dc_category"]),
        get_val(["GTIN", "gtin", "upc", "barcode"]),
        get_val(["Brand", "brand"]),
    ]


# =============================================================================
# DATABASE EXPORT FUNCTIONS
# =============================================================================

def export_count_session_for_upload(session_id: str) -> BytesIO:
    """
    Export a count session in Inventory Upload Template format.

    This produces a file that can be uploaded to OrderMaestro to
    update inventory counts.

    Args:
        session_id: Count session identifier

    Returns:
        BytesIO buffer containing the Excel file
    """
    import json
    from backend.core.database import (
        get_count_session, list_count_items, list_files, FileStatus
    )

    session = get_count_session(session_id)
    if not session:
        raise ValueError(f"Count session not found: {session_id}")

    site_id = session["site_id"]
    count_items = list_count_items(session_id)

    if not count_items:
        return create_inventory_upload_workbook(items=[])

    # Get original file data for additional columns
    files = list_files(status=FileStatus.COMPLETED, site_id=site_id, limit=1)
    original_rows = {}

    if files:
        file_record = files[0]
        parsed_data = file_record.get("parsed_data")
        if parsed_data:
            if isinstance(parsed_data, str):
                data = json.loads(parsed_data)
            else:
                data = parsed_data

            for row in data.get("rows", []):
                sku = row.get("Dist #") or row.get("sku") or ""
                if sku:
                    original_rows[str(sku).strip()] = row

    # Build export items
    export_items = []
    for count_item in count_items:
        sku = count_item.get("sku", "")

        # Start with original data if available
        if sku in original_rows:
            item = dict(original_rows[sku])
        else:
            item = {}

        # Override with count values
        item["Dist #"] = sku
        item["Item Description"] = count_item.get("description", "")
        item["Quantity"] = count_item.get("counted_qty", 0)
        item["UOM"] = count_item.get("uom", "")
        item["Price"] = count_item.get("unit_price")
        item["Location"] = count_item.get("location", "")

        export_items.append(item)

    return create_inventory_upload_workbook(items=export_items)


def export_cart_for_upload(site_id: str) -> BytesIO:
    """
    Export shopping cart in Cart Upload Template format.

    This produces a file that can be uploaded to OrderMaestro
    to populate a shopping cart.

    Args:
        site_id: Site identifier

    Returns:
        BytesIO buffer containing the Excel file
    """
    from backend.core.database import list_cart_items

    cart_items = list_cart_items(site_id)

    if not cart_items:
        return create_cart_upload_workbook(items=[])

    # Build export items - only need Dist # and Quantity
    export_items = []
    for cart_item in cart_items:
        export_items.append({
            "Dist #": cart_item.get("sku", ""),
            "GTIN": cart_item.get("gtin", ""),
            "Quantity": cart_item.get("quantity", 0),
            "Break Quantity": cart_item.get("break_quantity", ""),
        })

    return create_cart_upload_workbook(items=export_items)


def export_inventory_for_upload(site_id: str) -> BytesIO:
    """
    Export inventory in Inventory Upload Template format.

    This produces a file that can be uploaded to OrderMaestro
    to update inventory. Single sheet, 23 columns.

    Args:
        site_id: Site identifier

    Returns:
        BytesIO buffer containing the Excel file
    """
    import json
    from backend.core.database import list_files, FileStatus

    files = list_files(status=FileStatus.COMPLETED, site_id=site_id, limit=1)

    if not files:
        return create_inventory_upload_workbook(items=[])

    file_record = files[0]
    parsed_data = file_record.get("parsed_data")

    if not parsed_data:
        return create_inventory_upload_workbook(items=[])

    if isinstance(parsed_data, str):
        data = json.loads(parsed_data)
    else:
        data = parsed_data

    rows = data.get("rows", [])

    return create_inventory_upload_workbook(items=rows)


def export_inventory_as_valuation(site_id: str) -> BytesIO:
    """
    Export inventory in Valuation Report format (for records/archiving).

    This matches the format OrderMaestro produces when downloading
    a valuation report. Use this for record-keeping, not for uploads.

    Args:
        site_id: Site identifier

    Returns:
        BytesIO buffer containing the Excel file
    """
    import json
    from backend.core.database import (
        get_site_display_name, list_files, FileStatus
    )

    site_name = get_site_display_name(site_id)
    files = list_files(status=FileStatus.COMPLETED, site_id=site_id, limit=1)

    if not files:
        return create_valuation_report_workbook(site_name=site_name, items=[])

    file_record = files[0]
    parsed_data = file_record.get("parsed_data")

    if not parsed_data:
        return create_valuation_report_workbook(site_name=site_name, items=[])

    if isinstance(parsed_data, str):
        data = json.loads(parsed_data)
    else:
        data = parsed_data

    rows = data.get("rows", [])

    return create_valuation_report_workbook(
        site_name=site_name,
        items=rows
    )


def export_count_session_as_valuation(session_id: str) -> BytesIO:
    """
    Export count session in Valuation Report format (for records/archiving).

    Args:
        session_id: Count session identifier

    Returns:
        BytesIO buffer containing the Excel file
    """
    import json
    from backend.core.database import (
        get_count_session, get_site_display_name, list_count_items,
        list_files, FileStatus
    )

    session = get_count_session(session_id)
    if not session:
        raise ValueError(f"Count session not found: {session_id}")

    site_id = session["site_id"]
    site_name = get_site_display_name(site_id)
    count_items = list_count_items(session_id)

    if not count_items:
        return create_valuation_report_workbook(site_name=site_name, items=[])

    # Get original file data
    files = list_files(status=FileStatus.COMPLETED, site_id=site_id, limit=1)
    original_rows = {}

    if files:
        file_record = files[0]
        parsed_data = file_record.get("parsed_data")
        if parsed_data:
            if isinstance(parsed_data, str):
                data = json.loads(parsed_data)
            else:
                data = parsed_data

            for row in data.get("rows", []):
                sku = row.get("Dist #") or row.get("sku") or ""
                if sku:
                    original_rows[str(sku).strip()] = row

    # Build export items
    export_items = []
    for count_item in count_items:
        sku = count_item.get("sku", "")

        if sku in original_rows:
            item = dict(original_rows[sku])
        else:
            item = {}

        item["Dist #"] = sku
        item["Item Description"] = count_item.get("description", "")
        item["Quantity"] = count_item.get("counted_qty", 0)
        item["UOM"] = count_item.get("uom", "")
        item["Unit Price"] = count_item.get("unit_price")
        item["Compass Group USA->GL Codes"] = count_item.get("location", "")

        # Calculate total
        qty = count_item.get("counted_qty", 0) or 0
        price = count_item.get("unit_price") or 0
        item["Total Price"] = qty * price

        export_items.append(item)

    session_name = session.get("name", "Count")

    return create_valuation_report_workbook(
        site_name=site_name,
        items=export_items,
        printed_by=f"Spectre {session_name}"
    )


# =============================================================================
# LEGACY ALIASES (for backwards compatibility during refactor)
# =============================================================================

# These maintain backwards compatibility with existing code
def create_ordermaestro_workbook(site_name: str, items: List[Dict[str, Any]],
                                  printed_by: str = "Spectre") -> BytesIO:
    """Legacy alias for create_valuation_report_workbook."""
    return create_valuation_report_workbook(site_name, items, printed_by)

def export_inventory_from_db(site_id: str) -> BytesIO:
    """Legacy alias for export_inventory_as_valuation."""
    return export_inventory_as_valuation(site_id)

def export_cart_from_db(site_id: str) -> BytesIO:
    """Legacy - now exports in correct cart upload format."""
    return export_cart_for_upload(site_id)

def export_count_session_from_db(session_id: str) -> BytesIO:
    """Legacy - now exports in correct inventory upload format."""
    return export_count_session_for_upload(session_id)

"""
Unified Export Module - OrderMaestro Compatible Exports with Plugin Integration

Uses template-filling strategy to preserve exact OrderMaestro template structure.
Critical: "if you alter this template in any way, it will not upload"

Features:
- Template-based export (loads blank templates, fills data only)
- Off-catalog item integration
- Location-based sorting (walking order)
- Distributor validation and warnings
- Plugin-based categorization
"""

import json
import logging
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from backend.core.xlsx_export import (
    INVENTORY_UPLOAD_COLUMNS,
    CART_UPLOAD_COLUMNS,
    extract_inventory_upload_row,
    create_valuation_report_workbook,
)
from backend.core.template_filler import TemplateFiller

logger = logging.getLogger(__name__)

# Fallback walking order (used if no plugin loaded)
_DEFAULT_LOCATION_ORDER = {
    'Freezer': 1,
    'Walk In Cooler': 2,
    'Beverage Room': 3,
    'Dry Storage Food': 4,
    'Dry Storage Supplies': 5,
    'Chemical Locker': 6,
    'NEVER INVENTORY': 99,
    'UNASSIGNED': 100,
}


def _get_location_order() -> Dict[str, int]:
    """Get location order from plugin or use defaults."""
    try:
        from backend.core.plugins import PluginLoader
        loader = PluginLoader.get()
        if loader.has_plugins():
            plugin_order = loader.get_location_order()
            if plugin_order:
                return plugin_order
    except ImportError:
        pass
    return _DEFAULT_LOCATION_ORDER


def get_location_sort_key(location: str) -> int:
    """Get sort order for a location using plugin config."""
    if not location:
        return 100

    location_order = _get_location_order()

    # Try exact match first
    if location in location_order:
        return location_order[location]

    # Try case-insensitive match
    loc_lower = location.lower()
    for name, order in location_order.items():
        if name.lower() in loc_lower or loc_lower in name.lower():
            return order

    return 50  # Middle priority for unknown locations


def merge_off_catalog_items(
    inventory_items: List[Dict],
    site_id: str,
) -> List[Dict]:
    """
    Merge off-catalog items into inventory list.

    Off-catalog items are custom items not in the Master Order Guide.
    They're identified by having a Cust # but unusual Dist #.
    """
    from backend.core.database import list_off_catalog_items

    try:
        off_catalog = list_off_catalog_items(site_id, include_inactive=False)
    except Exception as e:
        logger.warning(f"Failed to load off-catalog items: {e}")
        return inventory_items

    if not off_catalog:
        return inventory_items

    # Create lookup of existing items by Dist #
    existing_skus = set()
    for item in inventory_items:
        sku = item.get('Dist #') or item.get('dist_num') or item.get('sku') or ''
        if sku:
            existing_skus.add(str(sku).strip().upper())

    # Add off-catalog items that aren't already in inventory
    merged = list(inventory_items)
    added_count = 0

    for oc_item in off_catalog:
        dist_num = oc_item.get('dist_num', '')
        if dist_num and str(dist_num).strip().upper() not in existing_skus:
            # Convert off-catalog item to inventory format
            merged.append({
                'Item Description': oc_item.get('description', ''),
                'Dist #': dist_num,
                'Cust #': oc_item.get('cust_num', ''),
                'Quantity': 0,  # Off-catalog items start with 0 qty
                'UOM': oc_item.get('uom', ''),
                'Pack': oc_item.get('pack', ''),
                'Price': oc_item.get('unit_price'),
                'Distributor': oc_item.get('distributor', ''),
                'Brand': oc_item.get('brand', ''),
                'GTIN': oc_item.get('gtin', ''),
                'Location': oc_item.get('location', ''),
                'Area': oc_item.get('area', ''),
                'Place': oc_item.get('place', ''),
                '_is_off_catalog': True,  # Internal marker
            })
            added_count += 1

    if added_count > 0:
        logger.info(f"Merged {added_count} off-catalog items into export")

    return merged


def categorize_items(items: List[Dict]) -> List[Dict]:
    """
    Apply location categorization to items using plugin rules.

    Only categorizes items that don't already have a location assigned.
    """
    try:
        from backend.core.plugins import PluginLoader
        loader = PluginLoader.get()
        if not loader.has_plugins():
            return items
    except ImportError:
        return items

    categorized = []
    for item in items:
        item_copy = dict(item)

        # Get current location
        location = (
            item_copy.get('Location') or
            item_copy.get('location') or
            item_copy.get('storage_location') or
            ''
        )

        # If no location, try to categorize
        if not location or location.upper() in ('UNASSIGNED', 'UNKNOWN', ''):
            desc = item_copy.get('Item Description') or item_copy.get('description') or ''
            brand = item_copy.get('Brand') or item_copy.get('brand') or ''
            pack = item_copy.get('Pack') or item_copy.get('pack') or ''

            new_location, never_count = loader.categorize_item(desc, brand, pack)
            item_copy['Location'] = new_location
            item_copy['_auto_categorized'] = True
            item_copy['_never_count'] = never_count

        categorized.append(item_copy)

    return categorized


def sort_by_location(items: List[Dict]) -> List[Dict]:
    """
    Sort items by location walking order, then by description.
    """
    def sort_key(item):
        location = (
            item.get('Location') or
            item.get('location') or
            'UNASSIGNED'
        )
        desc = (
            item.get('Item Description') or
            item.get('description') or
            ''
        ).upper()
        return (get_location_sort_key(location), desc)

    return sorted(items, key=sort_key)


def validate_distributors(items: List[Dict]) -> Tuple[List[Dict], List[Dict]]:
    """
    Validate distributors using plugin rules.

    Returns:
        (items_with_warnings, warning_list)
        - items_with_warnings: Items with _distributor_warning added if flagged
        - warning_list: List of warning dicts for flagged distributors
    """
    try:
        from backend.core.plugins import PluginLoader
        loader = PluginLoader.get()
        if not loader.has_plugins():
            return items, []
    except ImportError:
        return items, []

    warnings = []
    validated = []
    seen_distributors = set()

    for item in items:
        item_copy = dict(item)
        distributor = item_copy.get('Distributor') or item_copy.get('distributor') or ''

        if distributor and distributor not in seen_distributors:
            is_flagged, reason, severity = loader.is_distributor_flagged(distributor)

            if is_flagged:
                seen_distributors.add(distributor)
                item_copy['_distributor_warning'] = {
                    'reason': reason,
                    'severity': severity,
                }
                warnings.append({
                    'distributor': distributor,
                    'reason': reason,
                    'severity': severity,
                })

        validated.append(item_copy)

    return validated, warnings


def validate_gl_codes(items: List[Dict]) -> Tuple[List[Dict], List[Dict]]:
    """
    Validate GL code assignments on items.

    Flags:
    - Items with no GL code assigned
    - Items with multiple GL codes (comma-separated or multiple assignments)

    Returns:
        (items_with_flags, gl_issues_list)
    """
    GL_CODE_KEYS = [
        'Compass Group USA->GL Codes',
        'GL Codes',
        'gl_code',
        'GL Code',
    ]

    def get_gl_code(item: Dict) -> str:
        for key in GL_CODE_KEYS:
            if key in item and item[key]:
                return str(item[key]).strip()
        return ''

    issues = []
    validated = []

    for item in items:
        item_copy = dict(item)
        gl_code = get_gl_code(item_copy)
        desc = (
            item_copy.get('Item Description') or
            item_copy.get('description') or
            'Unknown'
        )[:50]

        if not gl_code:
            # No GL code assigned
            item_copy['_gl_code_issue'] = {
                'type': 'missing',
                'message': 'No GL code assigned',
                'severity': 'warning',
            }
            issues.append({
                'item': desc,
                'type': 'missing',
                'message': 'No GL code assigned',
                'severity': 'warning',
            })
        elif ',' in gl_code or gl_code.count('->') > 1:
            # Multiple GL codes detected
            item_copy['_gl_code_issue'] = {
                'type': 'multiple',
                'message': f'Multiple GL codes: {gl_code}',
                'severity': 'warning',
            }
            issues.append({
                'item': desc,
                'type': 'multiple',
                'gl_codes': gl_code,
                'message': 'Multiple GL codes assigned',
                'severity': 'warning',
            })

        validated.append(item_copy)

    return validated, issues


def create_unified_inventory_export(
    site_id: str,
    include_off_catalog: bool = True,
    sort_by_walking_order: bool = True,
    auto_categorize: bool = True,
    validate_distributor_flags: bool = True,
    validate_gl_codes_flag: bool = True,
    exclude_never_count: bool = False,
) -> Tuple[BytesIO, Dict[str, Any]]:
    """
    Create a unified inventory export with all enhancements.

    Args:
        site_id: Site identifier
        include_off_catalog: Merge off-catalog items into export
        sort_by_walking_order: Sort items by location walking order
        auto_categorize: Auto-categorize items without locations
        validate_distributor_flags: Check for flagged distributors
        validate_gl_codes_flag: Check for missing/multiple GL codes
        exclude_never_count: Exclude items marked as NEVER INVENTORY

    Returns:
        (BytesIO buffer, metadata dict with stats and warnings)
    """
    from backend.core.database import list_files, FileStatus

    # Get latest inventory data
    files = list_files(status=FileStatus.COMPLETED, site_id=site_id, limit=1)

    if not files:
        return _create_empty_workbook(), {'items': 0, 'warnings': []}

    file_record = files[0]
    parsed_data = file_record.get('parsed_data')

    if not parsed_data:
        return _create_empty_workbook(), {'items': 0, 'warnings': []}

    if isinstance(parsed_data, str):
        data = json.loads(parsed_data)
    else:
        data = parsed_data

    items = data.get('rows', [])
    metadata = {
        'original_count': len(items),
        'warnings': [],
        'gl_code_issues': [],
        'off_catalog_added': 0,
        'auto_categorized': 0,
        'excluded_never_count': 0,
    }

    # Step 1: Merge off-catalog items
    if include_off_catalog:
        before_count = len(items)
        items = merge_off_catalog_items(items, site_id)
        metadata['off_catalog_added'] = len(items) - before_count

    # Step 2: Auto-categorize
    if auto_categorize:
        items = categorize_items(items)
        metadata['auto_categorized'] = sum(
            1 for i in items if i.get('_auto_categorized')
        )

    # Step 3: Validate distributors
    if validate_distributor_flags:
        items, warnings = validate_distributors(items)
        metadata['warnings'] = warnings

    # Step 4: Validate GL codes
    if validate_gl_codes_flag:
        items, gl_issues = validate_gl_codes(items)
        metadata['gl_code_issues'] = gl_issues

    # Step 5: Exclude NEVER INVENTORY items if requested
    if exclude_never_count:
        before_count = len(items)
        items = [
            i for i in items
            if not i.get('_never_count') and
               i.get('Location', '').upper() != 'NEVER INVENTORY'
        ]
        metadata['excluded_never_count'] = before_count - len(items)

    # Step 6: Sort by location
    if sort_by_walking_order:
        items = sort_by_location(items)

    metadata['final_count'] = len(items)

    # Try to use template-filling strategy
    buffer = _create_inventory_from_template(site_id, items)

    if buffer is None:
        # Fallback to from-scratch creation if no template found
        logger.warning(f"No template found for {site_id}, creating from scratch")
        buffer = _create_inventory_from_scratch(items)

    return buffer, metadata


def _create_inventory_from_template(site_id: str, items: List[Dict]) -> Optional[BytesIO]:
    """
    Create inventory export using template-filling strategy.

    Loads the blank template and fills it with data, preserving
    all formatting and structure for OrderMaestro compatibility.
    """
    try:
        from backend.core.plugins import PluginLoader
        loader = PluginLoader.get()

        template_path = loader.get_inventory_template_path(site_id)
        if not template_path:
            return None

        logger.info(f"Using template: {template_path}")

        filler = TemplateFiller(template_path)
        return filler.fill_inventory(items)

    except ImportError:
        logger.warning("Plugin system not available")
        return None
    except Exception as e:
        logger.error(f"Template filling failed: {e}")
        return None


def _create_inventory_from_scratch(items: List[Dict]) -> BytesIO:
    """
    Fallback: Create inventory workbook from scratch.

    Only used when no template is available.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Upload"

    # Header row
    ws.append(INVENTORY_UPLOAD_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Highlight fills
    off_catalog_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    distributor_warning_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    gl_code_issue_fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")

    # Data rows
    for item in items:
        row_data = extract_inventory_upload_row(item)
        ws.append(row_data)

        row_num = ws.max_row
        if item.get('_distributor_warning'):
            for cell in ws[row_num]:
                cell.fill = distributor_warning_fill
        elif item.get('_gl_code_issue'):
            for cell in ws[row_num]:
                cell.fill = gl_code_issue_fill
        elif item.get('_is_off_catalog'):
            for cell in ws[row_num]:
                cell.fill = off_catalog_fill

    # Set column widths
    column_widths = [35, 12, 12, 10, 12, 8, 10, 20, 15, 15, 20, 15, 15, 12, 12, 18, 10, 10, 20, 15, 12, 12, 12]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _create_empty_workbook() -> BytesIO:
    """Create an empty inventory upload workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Upload"
    ws.append(INVENTORY_UPLOAD_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def create_unified_cart_export(
    site_id: str,
    validate_distributor_flags: bool = True,
) -> Tuple[BytesIO, Dict[str, Any]]:
    """
    Create a unified cart export with validation.

    Uses template-filling strategy when available.

    Args:
        site_id: Site identifier
        validate_distributor_flags: Check for flagged distributors

    Returns:
        (BytesIO buffer, metadata dict)
    """
    from backend.core.database import list_cart_items

    cart_items = list_cart_items(site_id)
    metadata = {
        'items': len(cart_items),
        'warnings': [],
    }

    if not cart_items:
        # Try to return empty template, fallback to from-scratch
        buffer = _create_cart_from_template([])
        if buffer is None:
            buffer = _create_cart_from_scratch([])
        return buffer, metadata

    # Validate distributors if requested
    if validate_distributor_flags:
        cart_items, warnings = validate_distributors(cart_items)
        metadata['warnings'] = warnings

    # Try template-filling strategy first
    buffer = _create_cart_from_template(cart_items)

    if buffer is None:
        # Fallback to from-scratch creation
        logger.warning("No cart template found, creating from scratch")
        buffer = _create_cart_from_scratch(cart_items)

    return buffer, metadata


def _create_cart_from_template(items: List[Dict]) -> Optional[BytesIO]:
    """
    Create cart export using template-filling strategy.

    Loads the blank cart template and fills it with data.
    """
    try:
        from backend.core.plugins import PluginLoader
        loader = PluginLoader.get()

        template_path = loader.get_cart_template_path()
        if not template_path:
            return None

        logger.info(f"Using cart template: {template_path}")

        filler = TemplateFiller(template_path)
        return filler.fill_cart(items)

    except ImportError:
        logger.warning("Plugin system not available for cart template")
        return None
    except Exception as e:
        logger.error(f"Cart template filling failed: {e}")
        return None


def _create_cart_from_scratch(items: List[Dict]) -> BytesIO:
    """
    Fallback: Create cart workbook from scratch.

    Only used when no template is available.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Shopping Cart Upload"

    ws.append(CART_UPLOAD_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for item in items:
        row_data = [
            item.get('Dist #') or item.get('sku') or item.get('dist_num') or '',
            item.get('GTIN') or item.get('gtin') or '',
            item.get('Quantity') or item.get('quantity') or 0,
            item.get('Break Quantity') or item.get('break_quantity') or '',
        ]
        ws.append(row_data)

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 15

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def validate_ordermaestro_format(items: List[Dict]) -> List[str]:
    """
    Validate items for OrderMaestro upload compatibility.

    Returns list of validation errors/warnings.
    """
    errors = []

    for idx, item in enumerate(items, 1):
        dist_num = item.get('Dist #') or item.get('dist_num') or ''

        # Dist # is required
        if not dist_num:
            desc = item.get('Item Description') or item.get('description') or 'Unknown'
            errors.append(f"Row {idx}: Missing Dist # for '{desc[:30]}'")

        # Quantity should be numeric
        qty = item.get('Quantity') or item.get('quantity')
        if qty is not None:
            try:
                float(qty)
            except (ValueError, TypeError):
                errors.append(f"Row {idx}: Invalid quantity '{qty}'")

    return errors

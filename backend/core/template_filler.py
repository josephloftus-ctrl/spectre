"""
Template Filler - OrderMaestro Compatible Template-Based Export

CRITICAL: This module fills existing BLANK templates downloaded from OrderMaestro.
The templates must remain EXACTLY as downloaded - we only write cell VALUES,
never modify structure, formatting, or metadata.

Key principle from training docs: "if you alter this template in any way, it will not upload"

Usage:
    from backend.core.template_filler import TemplateFiller

    filler = TemplateFiller(Path('plugins/culinart/templates/blank.xlsx'))
    buffer = filler.fill_inventory(items)
"""

import logging
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# =============================================================================
# FIXED COLUMN MAPPINGS (from OrderMaestro training docs)
# These are the EXACT column positions - never change them
# =============================================================================

INVENTORY_COLUMNS = {
    'Item Description': 1,   # A
    'Dist #': 2,             # B - REQUIRED
    'Cust #': 3,             # C - Required for off-catalog
    'Quantity': 4,           # D
    'Break Quantity': 5,     # E
    'UOM': 6,                # F
    'Break Uom': 7,          # G
    'Location': 8,           # H
    'Area': 9,               # I
    'Place': 10,             # J
    'Distribution Center': 11,  # K
    'Brand': 12,             # L
    'Mfg': 13,               # M
    'Mfg #': 14,             # N
    'Pack': 15,              # O
    'GTIN': 16,              # P
    'Price': 17,             # Q
    'Break Price': 18,       # R
    'Distributor': 19,       # S
    'Upc': 20,               # T
    'Catch Weight': 21,      # U
    'Average Weight': 22,    # V
    'Units Per Case': 23,    # W
}

CART_COLUMNS = {
    'Dist #': 1,             # A - REQUIRED
    'GTIN': 2,               # B
    'Quantity': 3,           # C - REQUIRED
    'Break Quantity': 4,     # D
}

SHOPPING_LIST_COLUMNS = {
    'Dist #': 1,             # A - REQUIRED
    'GTIN': 2,               # B
    'Cust #': 3,             # C
}

# Field name fallback mappings (our internal names -> template names)
# Note: Template headers may have asterisks for required fields (e.g., "Dist # *")
FIELD_FALLBACKS = {
    'Item Description': ['description', 'item_name', 'name'],
    'Dist #': ['Dist # *', 'dist_num', 'sku', 'item_number', 'dist_number'],
    'Cust #': ['Cust # *', 'cust_num', 'customer_number', 'cust_number'],
    'Quantity': ['quantity', 'qty', 'counted_qty', 'count'],
    'Break Quantity': ['break_quantity', 'break_qty'],
    'UOM': ['uom', 'unit_of_measure', 'unit'],
    'Break Uom': ['break_uom', 'break_unit'],
    'Location': ['location', 'storage_location', 'loc'],
    'Area': ['area', 'sub_location'],
    'Place': ['place', 'secondary_location'],
    'Distribution Center': ['DC Name', 'dc_name', 'distribution_center'],
    'Brand': ['brand'],
    'Mfg': ['mfg', 'manufacturer'],
    'Mfg #': ['mfg_num', 'mfg_number', 'manufacturer_number'],
    'Pack': ['Pack Type', 'pack', 'pack_size'],  # Template uses "Pack Type"
    'GTIN': ['gtin', 'barcode', 'upc_code'],
    'Price': ['Unit Price', 'unit_price', 'price'],
    'Break Price': ['break_price'],
    'Distributor': ['distributor', 'vendor'],
    'Upc': ['upc'],
    'Catch Weight': ['catch_weight'],
    'Average Weight': ['average_weight'],
    'Units Per Case': ['units_per_case'],
}

# Columns that should be numeric (for type coercion)
NUMERIC_COLUMNS = {
    'Quantity', 'Break Quantity', 'Price', 'Break Price',
    'Catch Weight', 'Average Weight', 'Units Per Case'
}


class TemplateFiller:
    """
    Fills OrderMaestro templates with data.

    IMPORTANT: Only writes cell values. Never modifies:
    - Row 1 (headers)
    - Cell formatting/styles
    - Sheet structure
    - Named ranges
    - Data validation
    - Any other metadata
    """

    def __init__(self, template_path: Path):
        """
        Initialize with path to template file.

        Args:
            template_path: Path to the blank template file
        """
        self.template_path = template_path
        self.wb = None
        self._load_template()

    def _load_template(self):
        """Load template with all preservation flags."""
        try:
            self.wb = load_workbook(
                filename=str(self.template_path),
                keep_vba=True,       # Preserve macros if any
                data_only=False,     # Keep formulas
                keep_links=True,     # Keep external links
                rich_text=False,     # Don't need rich text
            )
            logger.debug(f"Loaded template: {self.template_path}")
        except Exception as e:
            logger.error(f"Failed to load template {self.template_path}: {e}")
            raise

    def fill_inventory(self, items: List[Dict[str, Any]]) -> BytesIO:
        """
        Fill inventory template with items.

        Args:
            items: List of inventory items with standard field names

        Returns:
            BytesIO buffer containing filled template
        """
        ws = self.wb.active

        # Data starts at row 2 (row 1 is headers - DO NOT TOUCH)
        for row_idx, item in enumerate(items, start=2):
            self._write_row(ws, row_idx, item, INVENTORY_COLUMNS)

        logger.info(f"Filled inventory template with {len(items)} items")
        return self._save_to_buffer()

    def fill_cart(self, items: List[Dict[str, Any]]) -> BytesIO:
        """
        Fill shopping cart template with items.

        Args:
            items: List of cart items with sku and quantity

        Returns:
            BytesIO buffer containing filled template
        """
        ws = self.wb.active

        for row_idx, item in enumerate(items, start=2):
            self._write_row(ws, row_idx, item, CART_COLUMNS)

        logger.info(f"Filled cart template with {len(items)} items")
        return self._save_to_buffer()

    def fill_shopping_list(self, items: List[Dict[str, Any]]) -> BytesIO:
        """
        Fill shopping list template with items.

        Args:
            items: List of items with item numbers

        Returns:
            BytesIO buffer containing filled template
        """
        ws = self.wb.active

        for row_idx, item in enumerate(items, start=2):
            self._write_row(ws, row_idx, item, SHOPPING_LIST_COLUMNS)

        logger.info(f"Filled shopping list template with {len(items)} items")
        return self._save_to_buffer()

    def _write_row(
        self,
        ws,
        row_idx: int,
        item: Dict[str, Any],
        column_map: Dict[str, int]
    ):
        """
        Write a single row - VALUES ONLY.

        Never modifies cell formatting, only sets values.
        """
        for field_name, col_idx in column_map.items():
            value = self._get_field_value(item, field_name)
            if value is not None:
                # Coerce to appropriate type
                coerced_value = self._coerce_value(value, field_name)
                # Only set cell value - never touch formatting
                ws.cell(row=row_idx, column=col_idx).value = coerced_value

    def _get_field_value(self, item: Dict[str, Any], field_name: str) -> Any:
        """
        Get field value with fallback key mapping.

        Returns None if field not found (cell stays empty).
        """
        # Direct match first
        if field_name in item and item[field_name] is not None:
            return item[field_name]

        # Case-insensitive direct match
        for key, val in item.items():
            if key and key.lower() == field_name.lower() and val is not None:
                return val

        # Try fallback mappings
        for alt_key in FIELD_FALLBACKS.get(field_name, []):
            if alt_key in item and item[alt_key] is not None:
                return item[alt_key]
            # Case-insensitive fallback
            for key, val in item.items():
                if key and key.lower() == alt_key.lower() and val is not None:
                    return val

        return None

    def _coerce_value(self, value: Any, field_name: str) -> Any:
        """
        Force correct data types for OrderMaestro validation.

        Numeric fields should be actual numbers, not strings.
        """
        if value is None:
            return None

        # Numeric columns should be actual numbers
        if field_name in NUMERIC_COLUMNS:
            if isinstance(value, (int, float)):
                return value
            try:
                # Try to parse as number
                str_val = str(value).strip()
                if not str_val:
                    return None
                # Remove currency symbols and commas
                str_val = str_val.replace('$', '').replace(',', '')
                if '.' in str_val:
                    return float(str_val)
                return int(str_val)
            except (ValueError, TypeError):
                # If can't convert, return as-is
                return value

        # Text columns - return as string if not None
        if value is not None:
            return str(value) if not isinstance(value, str) else value

        return value

    def _save_to_buffer(self) -> BytesIO:
        """Save workbook to BytesIO buffer."""
        buffer = BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        return buffer


def fill_inventory_template(
    template_path: Path,
    items: List[Dict[str, Any]]
) -> BytesIO:
    """
    Convenience function to fill an inventory template.

    Args:
        template_path: Path to blank inventory template
        items: List of inventory items

    Returns:
        BytesIO buffer containing filled template
    """
    filler = TemplateFiller(template_path)
    return filler.fill_inventory(items)


def fill_cart_template(
    template_path: Path,
    items: List[Dict[str, Any]]
) -> BytesIO:
    """
    Convenience function to fill a cart template.

    Args:
        template_path: Path to blank cart template
        items: List of cart items

    Returns:
        BytesIO buffer containing filled template
    """
    filler = TemplateFiller(template_path)
    return filler.fill_cart(items)

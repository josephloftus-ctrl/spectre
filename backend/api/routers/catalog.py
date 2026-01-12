"""
Off-catalog items API router.
"""
from fastapi import APIRouter, HTTPException, UploadFile, File, Form, Query
from io import BytesIO
import openpyxl

from backend.core.database import (
    create_off_catalog_item, get_off_catalog_item, get_off_catalog_item_by_dist,
    update_off_catalog_item, delete_off_catalog_item, list_off_catalog_items,
    bulk_import_off_catalog_items, generate_cust_num
)
from backend.api.models import OffCatalogItemRequest

router = APIRouter(prefix="/api/off-catalog", tags=["Off-Catalog"])


@router.get("/{site_id}")
def list_site_off_catalog_items(
    site_id: str,
    include_inactive: bool = Query(False)
):
    """List all off-catalog items for a site."""
    items = list_off_catalog_items(site_id, include_inactive=include_inactive)
    return {"items": items, "count": len(items)}


@router.get("/{site_id}/{cust_num}")
def get_off_catalog_by_cust(site_id: str, cust_num: str):
    """Get an off-catalog item by customer number."""
    item = get_off_catalog_item(site_id, cust_num)
    if not item:
        raise HTTPException(status_code=404, detail="Off-catalog item not found")
    return {"item": item}


@router.get("/{site_id}/by-dist/{dist_num}")
def get_off_catalog_by_dist_num(site_id: str, dist_num: str):
    """Get an off-catalog item by distributor number."""
    item = get_off_catalog_item_by_dist(site_id, dist_num)
    if not item:
        raise HTTPException(status_code=404, detail="Off-catalog item not found")
    return {"item": item}


@router.post("/{site_id}")
def create_new_off_catalog_item(site_id: str, request: OffCatalogItemRequest):
    """
    Create a new off-catalog item.

    Off-catalog items are custom items not in the Master Order Guide.
    They require a Dist # (for ordering) and Cust # (for barcode scanning).
    If Cust # is not provided, one will be auto-generated.
    """
    cust_num = request.cust_num
    if not cust_num:
        cust_num = generate_cust_num(site_id)

    try:
        item = create_off_catalog_item(
            site_id=site_id,
            dist_num=request.dist_num,
            cust_num=cust_num,
            description=request.description,
            pack=request.pack,
            uom=request.uom,
            unit_price=request.unit_price,
            distributor=request.distributor,
            break_uom=request.break_uom,
            break_price=request.break_price,
            distribution_center=request.distribution_center,
            brand=request.brand,
            manufacturer=request.manufacturer,
            manufacturer_num=request.manufacturer_num,
            gtin=request.gtin,
            upc=request.upc,
            catch_weight=request.catch_weight,
            average_weight=request.average_weight,
            units_per_case=request.units_per_case,
            location=request.location,
            area=request.area,
            place=request.place,
            notes=request.notes
        )
        return {"success": True, "item": item}
    except Exception as e:
        if "UNIQUE constraint failed" in str(e):
            raise HTTPException(
                status_code=409,
                detail=f"Off-catalog item with Cust # '{cust_num}' already exists"
            )
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/{site_id}/{cust_num}")
def update_off_catalog(site_id: str, cust_num: str, request: OffCatalogItemRequest):
    """Update an existing off-catalog item."""
    existing = get_off_catalog_item(site_id, cust_num)
    if not existing:
        raise HTTPException(status_code=404, detail="Off-catalog item not found")

    kwargs = {}
    for field, value in request.model_dump().items():
        if value is not None and field != "cust_num":
            kwargs[field] = value

    item = update_off_catalog_item(site_id, cust_num, **kwargs)
    return {"success": True, "item": item}


@router.delete("/{site_id}/{cust_num}")
def delete_off_catalog(
    site_id: str,
    cust_num: str,
    hard_delete: bool = Query(False)
):
    """
    Delete an off-catalog item.

    By default, soft-deletes (marks inactive). Use hard_delete=true
    to permanently remove the item.
    """
    success = delete_off_catalog_item(site_id, cust_num, hard_delete=hard_delete)
    if not success:
        raise HTTPException(status_code=404, detail="Off-catalog item not found")
    return {"success": True, "message": f"Deleted off-catalog item {cust_num}"}


@router.post("/{site_id}/bulk")
async def bulk_import_off_catalog(
    site_id: str,
    file: UploadFile = File(...),
    update_existing: bool = Form(True)
):
    """
    Bulk import off-catalog items from an Excel file.

    Expected columns: Dist #, Cust #, Item Description, Pack, UOM, Price, etc.
    If Cust # is missing for a row, one will be auto-generated.
    """
    contents = await file.read()

    try:
        wb = openpyxl.load_workbook(BytesIO(contents), data_only=True)
        ws = wb.active

        headers = [cell.value for cell in ws[1] if cell.value]
        if not headers:
            raise HTTPException(status_code=400, detail="No headers found in Excel file")

        header_map = {
            "Dist #": "dist_num",
            "Cust #": "cust_num",
            "Item Description": "description",
            "Pack": "pack",
            "UOM": "uom",
            "Break Uom": "break_uom",
            "Price": "unit_price",
            "Unit Price": "unit_price",
            "Break Price": "break_price",
            "Distributor": "distributor",
            "Distribution Center": "distribution_center",
            "Brand": "brand",
            "Mfg": "manufacturer",
            "Manufacturer": "manufacturer",
            "Mfg #": "manufacturer_num",
            "GTIN": "gtin",
            "Upc": "upc",
            "UPC": "upc",
            "Catch Weight": "catch_weight",
            "Average Weight": "average_weight",
            "Units Per Case": "units_per_case",
            "Location": "location",
            "Area": "area",
            "Place": "place",
            "Notes": "notes",
        }

        items = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue

            item = {}
            for col_idx, value in enumerate(row):
                if col_idx < len(headers):
                    header = headers[col_idx]
                    field = header_map.get(header, header.lower().replace(" ", "_"))
                    item[field] = value

            if not item.get("cust_num"):
                item["cust_num"] = generate_cust_num(site_id)

            items.append(item)

        if not items:
            raise HTTPException(status_code=400, detail="No data rows found in Excel file")

        results = bulk_import_off_catalog_items(site_id, items, update_existing=update_existing)
        return {
            "success": True,
            "results": results,
            "message": f"Imported {results['created']} new, updated {results['updated']}, skipped {results['skipped']}"
        }

    except Exception as e:
        if isinstance(e, HTTPException):
            raise
        raise HTTPException(status_code=400, detail=f"Failed to process file: {str(e)}")


@router.post("/{site_id}/generate-cust-num")
def generate_new_cust_num(site_id: str, prefix: str = Query("SPEC")):
    """Generate a new unique Cust # for off-catalog items."""
    cust_num = generate_cust_num(site_id, prefix=prefix)
    return {"cust_num": cust_num}

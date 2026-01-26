"""
Template Filler Test Suite

Tests that our template filler produces output compatible with MyOrders/OrderMaestro.
Uses the actual blank template as source of truth - we NEVER modify it.

Key requirements from training docs:
1. "If you alter this template in any way, it will not upload"
2. Off-catalog items require: Dist # + Cust #
3. On-MOG items require: Dist # only (+ optional Location/Area/Place)
"""

import pytest
from pathlib import Path
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Paths to templates (source of truth - never modify these)
TEMPLATES_DIR = Path(__file__).resolve().parents[2] / "Templates"
BLANK_TEMPLATE = TEMPLATES_DIR / "EmptyInventoryTemplate.xlsx"
FILLED_REFERENCE = TEMPLATES_DIR / "PSEG NHQ Inventory Template.xlsx"

# Import the module under test
from backend.core.template_filler import TemplateFiller, INVENTORY_COLUMNS


class TestTemplateStructure:
    """Test that we understand the template structure correctly."""

    def test_blank_template_exists(self):
        """Verify blank template file exists."""
        assert BLANK_TEMPLATE.exists(), f"Blank template not found at {BLANK_TEMPLATE}"

    def test_filled_reference_exists(self):
        """Verify filled reference template exists."""
        assert FILLED_REFERENCE.exists(), f"Filled reference not found at {FILLED_REFERENCE}"

    def test_blank_template_headers(self):
        """Verify we have the exact expected headers in correct positions."""
        wb = load_workbook(BLANK_TEMPLATE)
        ws = wb.active

        expected_headers = [
            (1, 'Item Description'),
            (2, 'Dist # *'),
            (3, 'Cust # *'),
            (4, 'Quantity'),
            (5, 'Break Quantity'),
            (6, 'UOM'),
            (7, 'Break Uom'),
            (8, 'Location'),
            (9, 'Area'),
            (10, 'Place'),
            (11, 'Distribution Center'),
            (12, 'Brand'),
            (13, 'Mfg'),
            (14, 'Mfg #'),
            (15, 'Pack Type'),  # Note: template says "Pack Type", not "Pack"
            (16, 'GTIN'),
            (17, 'Price'),
            (18, 'Break Price'),
            (19, 'Distributor'),
            (20, 'Upc'),
            (21, 'Catch Weight'),
            (22, 'Average Weight'),
            (23, 'Units Per Case'),
        ]

        for col_idx, expected_name in expected_headers:
            actual = ws.cell(row=1, column=col_idx).value
            assert actual == expected_name, (
                f"Column {col_idx} ({get_column_letter(col_idx)}): "
                f"expected '{expected_name}', got '{actual}'"
            )

    def test_column_mapping_matches_template(self):
        """Verify INVENTORY_COLUMNS mapping matches actual template positions."""
        wb = load_workbook(BLANK_TEMPLATE)
        ws = wb.active

        # Build a map of header name -> column index from the actual template
        actual_positions = {}
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header:
                # Strip asterisks for comparison
                clean_name = header.replace(' *', '').strip()
                actual_positions[clean_name] = col

        # Check our mapping
        for our_name, our_col in INVENTORY_COLUMNS.items():
            # Handle the "Pack" vs "Pack Type" discrepancy
            lookup_name = our_name
            if our_name == 'Pack':
                lookup_name = 'Pack Type'
            # Handle "Dist #" vs "Dist # *"
            if lookup_name in actual_positions:
                assert actual_positions[lookup_name] == our_col, (
                    f"Column '{our_name}' mapped to {our_col}, "
                    f"but template has it at {actual_positions[lookup_name]}"
                )


class TestFilledTemplateAnalysis:
    """Analyze the reference filled template to understand expected output format."""

    def test_data_types_in_reference(self):
        """Document the actual data types in the reference filled template."""
        wb = load_workbook(FILLED_REFERENCE)
        ws = wb.active

        # Check a sample row
        row = 2
        type_report = {}
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            value = ws.cell(row=row, column=col).value
            type_report[header] = {
                'value': repr(value)[:50],
                'type': type(value).__name__
            }

        # All values in the reference should be strings (including prices with $)
        price_val = ws.cell(row=2, column=17).value  # Price column
        assert isinstance(price_val, str), f"Price should be string, got {type(price_val)}"
        assert '$' in str(price_val), f"Price should contain $, got {price_val}"

    def test_empty_cells_in_reference(self):
        """Check how empty cells are represented in reference template."""
        wb = load_workbook(FILLED_REFERENCE)
        ws = wb.active

        # Quantity column should be empty (users fill this in)
        qty_val = ws.cell(row=2, column=4).value  # Quantity
        # Empty cells appear as empty string '' in the reference
        assert qty_val == '' or qty_val is None, f"Quantity should be empty, got {repr(qty_val)}"


class TestTemplateFiller:
    """Test the actual template filler functionality."""

    @pytest.fixture
    def filler(self):
        """Create a fresh TemplateFiller instance."""
        return TemplateFiller(BLANK_TEMPLATE)

    def test_headers_preserved_after_fill(self, filler):
        """Verify row 1 headers are NEVER modified."""
        # Get original headers
        original_wb = load_workbook(BLANK_TEMPLATE)
        original_headers = [
            original_wb.active.cell(row=1, column=c).value
            for c in range(1, 24)
        ]

        # Fill with test data
        test_items = [
            {'description': 'Test Item', 'sku': '123456', 'location': 'Freezer'}
        ]
        buffer = filler.fill_inventory(test_items)

        # Check headers in output
        output_wb = load_workbook(buffer)
        output_headers = [
            output_wb.active.cell(row=1, column=c).value
            for c in range(1, 24)
        ]

        assert original_headers == output_headers, "Headers were modified!"

    def test_data_written_to_correct_columns(self, filler):
        """Verify data goes to the right columns."""
        test_items = [{
            'description': 'RED BULL ENERGY DRINK',
            'sku': '2916452',
            'location': 'Beverage Room',
            'uom': 'CS',
        }]

        buffer = filler.fill_inventory(test_items)
        wb = load_workbook(buffer)
        ws = wb.active

        # Check row 2 (first data row)
        assert ws.cell(row=2, column=1).value == 'RED BULL ENERGY DRINK'  # Item Description
        assert ws.cell(row=2, column=2).value == '2916452'  # Dist #
        assert ws.cell(row=2, column=8).value == 'Beverage Room'  # Location
        assert ws.cell(row=2, column=6).value == 'CS'  # UOM

    def test_on_mog_item_minimal_fields(self, filler):
        """Test on-MOG items only need Dist # and Location."""
        # Per training doc: "You do not need to give on MOG items a Cust #, only the Dist #"
        test_items = [{
            'sku': '9351149',  # Dist #
            'location': 'Walk In',
        }]

        buffer = filler.fill_inventory(test_items)
        wb = load_workbook(buffer)
        ws = wb.active

        # Dist # should be populated
        assert ws.cell(row=2, column=2).value == '9351149'
        # Location should be populated
        assert ws.cell(row=2, column=8).value == 'Walk In'
        # Cust # should be empty (not required for on-MOG)
        cust_val = ws.cell(row=2, column=3).value
        assert cust_val is None or cust_val == '', f"Cust # should be empty for on-MOG, got {repr(cust_val)}"

    def test_off_catalog_item_requires_cust_num(self, filler):
        """Test off-catalog items need both Dist # and Cust #."""
        test_items = [{
            'description': 'Red Delicious Apples',
            'sku': '1234',  # Dist #
            'cust_num': '10001',  # Required for off-catalog
            'uom': 'CS',
            'location': 'Walk In',
            'distribution_center': "Joe's Produce",
        }]

        buffer = filler.fill_inventory(test_items)
        wb = load_workbook(buffer)
        ws = wb.active

        assert ws.cell(row=2, column=1).value == 'Red Delicious Apples'
        assert ws.cell(row=2, column=2).value == '1234'  # Dist #
        assert ws.cell(row=2, column=3).value == '10001'  # Cust #
        assert ws.cell(row=2, column=6).value == 'CS'

    def test_multiple_items(self, filler):
        """Test filling multiple rows."""
        test_items = [
            {'sku': '111', 'location': 'Freezer'},
            {'sku': '222', 'location': 'Cooler'},
            {'sku': '333', 'location': 'Dry Storage'},
        ]

        buffer = filler.fill_inventory(test_items)
        wb = load_workbook(buffer)
        ws = wb.active

        assert ws.cell(row=2, column=2).value == '111'
        assert ws.cell(row=3, column=2).value == '222'
        assert ws.cell(row=4, column=2).value == '333'

    def test_empty_template_unchanged_structure(self, filler):
        """Verify template structure (dimensions, etc.) is preserved."""
        original_wb = load_workbook(BLANK_TEMPLATE)
        original_ws = original_wb.active

        # Get original column widths
        original_widths = {
            col: dim.width
            for col, dim in original_ws.column_dimensions.items()
            if dim.width
        }

        # Fill and check
        test_items = [{'sku': '123', 'location': 'Test'}]
        buffer = filler.fill_inventory(test_items)
        output_wb = load_workbook(buffer)
        output_ws = output_wb.active

        # Column widths should match
        for col, width in original_widths.items():
            output_width = output_ws.column_dimensions[col].width
            assert output_width == width, (
                f"Column {col} width changed from {width} to {output_width}"
            )


class TestDataTypeCoercion:
    """Test that data types are handled correctly for MyOrders compatibility."""

    @pytest.fixture
    def filler(self):
        return TemplateFiller(BLANK_TEMPLATE)

    def test_price_format_preserved(self, filler):
        """
        Prices in the reference template are strings with $ symbol.
        We must preserve this format exactly - no stripping of $ or conversion to numbers.
        """
        test_items = [{
            'sku': '123',
            'price': '$53.35',  # Already formatted
        }]

        buffer = filler.fill_inventory(test_items)
        wb = load_workbook(buffer)
        ws = wb.active

        price = ws.cell(row=2, column=17).value
        # MUST preserve the string exactly as provided
        assert price == '$53.35', f"Price should be preserved as '$53.35', got {repr(price)}"

    def test_numeric_price_input(self, filler):
        """
        Test handling of numeric price input.
        Numeric values should be converted to string representation.
        """
        test_items = [{
            'sku': '123',
            'price': 53.35,  # Numeric input
        }]

        buffer = filler.fill_inventory(test_items)
        wb = load_workbook(buffer)
        ws = wb.active

        price = ws.cell(row=2, column=17).value
        # Numeric input gets converted to string
        assert price == '53.35', f"Numeric price should become string '53.35', got {repr(price)}"
        assert isinstance(price, str), f"Price should be string type, got {type(price).__name__}"

    def test_sku_as_string(self, filler):
        """SKUs should always be strings (may have leading zeros)."""
        test_items = [{
            'sku': '0012345',  # Leading zero
        }]

        buffer = filler.fill_inventory(test_items)
        wb = load_workbook(buffer)
        ws = wb.active

        sku = ws.cell(row=2, column=2).value
        assert sku == '0012345', f"SKU with leading zero: expected '0012345', got {repr(sku)}"

    def test_string_numeric_values_preserved(self, filler):
        """
        String numeric values like '0.00' and '24' should be preserved exactly.
        Reference template stores these as strings, not native numbers.
        """
        test_items = [{
            'sku': '123',
            'Average Weight': '0.00',
            'Units Per Case': '24',
            'Break Price': '$2.22',
        }]

        buffer = filler.fill_inventory(test_items)
        wb = load_workbook(buffer)
        ws = wb.active

        # All values should remain as exact strings
        avg_weight = ws.cell(row=2, column=22).value  # Average Weight
        units_per_case = ws.cell(row=2, column=23).value  # Units Per Case
        break_price = ws.cell(row=2, column=18).value  # Break Price

        assert avg_weight == '0.00', f"Average Weight should be '0.00', got {repr(avg_weight)}"
        assert units_per_case == '24', f"Units Per Case should be '24', got {repr(units_per_case)}"
        assert break_price == '$2.22', f"Break Price should be '$2.22', got {repr(break_price)}"


class TestFieldMapping:
    """Test the field name fallback mapping."""

    @pytest.fixture
    def filler(self):
        return TemplateFiller(BLANK_TEMPLATE)

    def test_various_field_names(self, filler):
        """Test that various field name formats map correctly."""
        test_cases = [
            # (input_field, input_value, expected_column, expected_value)
            ({'description': 'Test'}, 1, 'Test'),
            ({'Item Description': 'Test'}, 1, 'Test'),
            ({'sku': '123'}, 2, '123'),
            ({'Dist #': '123'}, 2, '123'),
            ({'dist_num': '123'}, 2, '123'),
            ({'location': 'Freezer'}, 8, 'Freezer'),
            ({'Location': 'Freezer'}, 8, 'Freezer'),
        ]

        for item_dict, expected_col, expected_val in test_cases:
            filler_instance = TemplateFiller(BLANK_TEMPLATE)
            buffer = filler_instance.fill_inventory([item_dict])
            wb = load_workbook(buffer)
            ws = wb.active

            actual = ws.cell(row=2, column=expected_col).value
            assert actual == expected_val, (
                f"Input {item_dict} -> column {expected_col}: "
                f"expected {repr(expected_val)}, got {repr(actual)}"
            )


class TestRegressionAgainstReference:
    """Compare our output against the known-good reference template."""

    def test_compare_sample_row_to_reference(self):
        """
        Create output that should match reference row 2,
        then compare cell-by-cell.
        """
        # Data from reference template row 2
        reference_item = {
            'Item Description': 'DRINK ENERGY',
            'Dist #': '2916452',
            'UOM': 'CS',
            'Break Uom': 'EA',
            'Location': 'Beverage Room',
            'Distribution Center': 'SYSCO PHILADELPHIA',
            'Brand': 'REDBULL',
            'Mfg': 'RED BULL NORTH AMERICA INC',
            'Mfg #': 'RB4816',
            'Pack': '24/12 OZ',
            'GTIN': '00611269917475',
            'Price': '$53.35',
            'Break Price': '$2.22',
            'Distributor': 'Sysco Corporation',
            'Average Weight': '0.00',
            'Units Per Case': '24',
        }

        # Generate output
        filler = TemplateFiller(BLANK_TEMPLATE)
        buffer = filler.fill_inventory([reference_item])

        # Load both for comparison
        output_wb = load_workbook(buffer)
        output_ws = output_wb.active

        ref_wb = load_workbook(FILLED_REFERENCE)
        ref_ws = ref_wb.active

        # Compare populated columns
        columns_to_check = [1, 2, 6, 7, 8, 11, 12, 13, 14, 15, 16, 17, 18, 19, 22, 23]

        mismatches = []
        for col in columns_to_check:
            ref_val = ref_ws.cell(row=2, column=col).value
            out_val = output_ws.cell(row=2, column=col).value

            if ref_val != out_val:
                header = ref_ws.cell(row=1, column=col).value
                mismatches.append({
                    'column': col,
                    'header': header,
                    'reference': repr(ref_val),
                    'output': repr(out_val),
                })

        if mismatches:
            msg = "Mismatches found:\n"
            for m in mismatches:
                msg += f"  Col {m['column']} ({m['header']}): ref={m['reference']}, out={m['output']}\n"
            pytest.fail(msg)


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
